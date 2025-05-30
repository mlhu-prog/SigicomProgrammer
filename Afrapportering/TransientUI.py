# -*- coding: utf-8 -*-

################################################################################
## Form generated from reading UI file 'TransientUI.ui'
##
## Created by: Qt User Interface Compiler version 6.8.2
##
## WARNING! All changes made in this file will be lost when recompiling UI file!
################################################################################

from PySide6.QtCore import (QCoreApplication, QDate, QDateTime, QLocale,
    QMetaObject, QObject, QPoint, QRect,
    QSize, QTime, QUrl, Qt)
from PySide6.QtGui import (QBrush, QColor, QConicalGradient, QCursor,
    QFont, QFontDatabase, QGradient, QIcon,
    QImage, QKeySequence, QLinearGradient, QPainter,
    QPalette, QPixmap, QRadialGradient, QTransform)
from PySide6.QtWidgets import (QApplication, QCheckBox, QFrame, QGridLayout,
    QHBoxLayout, QLabel, QLayout, QMainWindow, QPushButton,
    QScrollArea, QSizePolicy, QSplitter, QSpacerItem, QStatusBar,
    QTabWidget, QVBoxLayout, QWidget, QMessageBox, QDialog)

import os

from basic_functions import load_json
from basic_functions import save_json
from basic_functions import search_json
from basic_functions import calculate_vibration_data

from logger import get_logger

class TransientUI(QDialog):
    def __init__(self, project_id, atr, name, cache_main_dir, output_folder, transient_data_folder, parent=None):
        super().__init__(parent)

        self.project_id = project_id
        self.atr = atr
        self.name = name
        self.cache_main_dir = cache_main_dir
        self.output_folder = output_folder
        self.transient_data_folder = transient_data_folder

        self.logger = get_logger()

        self.initial_settings()

        self.setupUi(self)

    def initial_settings(self):
        filenames = ["trans_dict_final.json", "trans_dict.json"]
        
        self.trans_dict_final_path, self.trans_dict_path = [os.path.join(self.cache_main_dir, "Results", file) for file in filenames]

        self.trans_dict = load_json(self.trans_dict_path)
        
        self.info_dict = load_json(os.path.join(self.cache_main_dir, "API", "project_sensor_dict_final.json"))
        
    def setupUi(self, Dialog):
        if not Dialog.objectName():
            Dialog.setObjectName(u"Dialog")
        Dialog.resize(1120, 600)

        self.mainLayout = QVBoxLayout(Dialog)
        self.mainLayout.setObjectName(u"mainLayout")
        
        #self.verticalLayout_2 = QVBoxLayout(self.centralwidget)
        #self.verticalLayout_2.setObjectName(u"verticalLayout_2")

        #self.verticalLayout = QVBoxLayout()
        #self.verticalLayout.setObjectName(u"verticalLayout")

        self.verticalLayout_3 = QVBoxLayout()
        self.verticalLayout_3.setObjectName(u"verticalLayout_3")

        self.tab_setup(Dialog)

        self.horizontalLayout = QHBoxLayout()
        self.horizontalLayout.setSpacing(0)
        self.horizontalLayout.setObjectName(u"horizontalLayout")
        self.horizontalLayout.setContentsMargins(-1, 9, -1, -1)

        self.cancelButton = QPushButton(Dialog)
        self.cancelButton.setObjectName(u"pushButton_4")
        sizePolicy4 = QSizePolicy(QSizePolicy.Policy.Maximum, QSizePolicy.Policy.Fixed)
        sizePolicy4.setHorizontalStretch(0)
        sizePolicy4.setVerticalStretch(0)
        sizePolicy4.setHeightForWidth(self.cancelButton.sizePolicy().hasHeightForWidth())
        self.cancelButton.setSizePolicy(sizePolicy4)
        self.cancelButton.setMinimumSize(QSize(75, 0))
        self.cancelButton.clicked.connect(self.cancel_MainWindow)

        self.horizontalLayout.addWidget(self.cancelButton)

        self.hyperLink = QLabel(Dialog)
        self.hyperLink.setObjectName(u"label_4")
        
        # Set the text of the label to be a hyperlink
        self.hyperLink.setText('<a href="https://cowi.sharepoint.com/:w:/r/sites/A004371-project/Shared%20Documents/3%20Projekt%20dokumenter/Manualer%20og%20fremgangsm%C3%A5der/Analyse%20af%20transienter.docx?d=w64aa320c1af94608bc59b652f241242b&csf=1&web=1&e=2sUslB">Guide: Analyse af transienter</a>')
        self.hyperLink.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.hyperLink.setOpenExternalLinks(True)  # Enable opening links in browser

        self.horizontalLayout.addWidget(self.hyperLink)

        self.continueButton = QPushButton(Dialog)
        self.continueButton.setObjectName(u"pushButton")
        sizePolicy4.setHeightForWidth(self.continueButton.sizePolicy().hasHeightForWidth())
        self.continueButton.setSizePolicy(sizePolicy4)
        self.continueButton.setMinimumSize(QSize(75, 0))
        self.continueButton.clicked.connect(self.remove_checked_transients)

        self.horizontalLayout.addWidget(self.continueButton)
        
        self.verticalLayout_3.addLayout(self.horizontalLayout)
        

        self.mainLayout.addLayout(self.verticalLayout_3)

        self.retranslateUi(Dialog)

        self.tabWidget.setCurrentIndex(0)


        QMetaObject.connectSlotsByName(Dialog)

    def tab_setup(self, Dialog):

        self.tabs = {}
        self.selected_vars = {}

        self.tabWidget = QTabWidget(Dialog)
        self.tabWidget.setObjectName(u"tabWidget")        
        #sizePolicy = QSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Maximum)
        #sizePolicy.setHorizontalStretch(0)
        #sizePolicy.setVerticalStretch(0)
        #sizePolicy.setHeightForWidth(self.tabWidget.sizePolicy().hasHeightForWidth())
        #self.tabWidget.setSizePolicy(sizePolicy)
        #self.tabWidget.setMaximumSize(QSize(1600000, 800))

        for i, (adresse, data) in enumerate(self.trans_dict.items()):
            errors_list = data.get("Transienter", [])
            errors_list = {entry['timestamp']: entry for entry in errors_list}
            sensor_id = str(self.info_dict[adresse]["sensor_id"])
            self.trans_dict[adresse]["sensor_id"] = sensor_id
                
            transients = search_json(errors_list, sensor_id)

            tab = QWidget()
            tab.setObjectName(u"tab")
            self.tabWidget.addTab(tab, QCoreApplication.translate("Dialog", adresse, None))
            
            self.selected_vars[adresse] = {}

            if errors_list:

                gridLayout_2 = QGridLayout(tab)
                gridLayout_2.setObjectName(u"gridLayout_2")

                label = QLabel(tab)
                label.setObjectName(u"label")
                label.setText(QCoreApplication.translate("Dialog", 
                                                         u"<html><head/><body><p align=\"center\"><span style=\" font-size:14pt; font-weight:700;\">Transienter for {}, {}.</span></p></body></html>".format(adresse, sensor_id), None))

                gridLayout_2.addWidget(label, 0, 0, 1, 1)


                splitter = QSplitter(tab)
                splitter.setObjectName(u"splitter")
                splitter.setOrientation(Qt.Orientation.Vertical)
                label_2 = QLabel(splitter)
                label_2.setObjectName(u"label_2")
                label_2.setEnabled(True)
                label_2.setMaximumSize(QSize(16777215, 25))
                label_2.setText(QCoreApplication.translate("Dialog", u"<html><head/><body><p><span style=\" font-size:12pt; font-weight:700;\">V\u00e6lg alle:</span></p></body></html>", None))
        
                splitter.addWidget(label_2)

                gridLayout_2.addWidget(splitter, 2, 0, 1, 1)
                
                button_selectAll_layout = QHBoxLayout()
                button_selectAll_layout.setObjectName(u"horizontalLayout_2")
                button_selectAll_layout.setContentsMargins(-1, -1, -1, 0)

                gridLayout_2.addLayout(button_selectAll_layout, 3, 0, 1, 1)

                reel_button_all = QPushButton(tab)
                reel_button_all.setObjectName(u"checkBox_3")        
                reel_button_all.setText(QCoreApplication.translate("Dialog", u"Reel", None))
                reel_button_all.clicked.connect(lambda checked, tab_idx=i, a=adresse, value=1: self.check_all(tab_idx, a, value))

                button_selectAll_layout.addWidget(reel_button_all)

                slag_button_all = QPushButton(tab)
                slag_button_all.setObjectName(u"checkBox_2")
                slag_button_all.setText(QCoreApplication.translate("Dialog", u"Slag", None))
                slag_button_all.clicked.connect(lambda checked, tab_idx=i, a=adresse, value=2: self.check_all(tab_idx, a, value))

                button_selectAll_layout.addWidget(slag_button_all)

                arb_button_all = QPushButton(tab)
                arb_button_all.setObjectName(u"checkBox")
                arb_button_all.setText(QCoreApplication.translate("Dialog", u"Arbejde t\u00e6t p\u00e5 m\u00e5leren", None))
                arb_button_all.clicked.connect(lambda checked, tab_idx=i, a=adresse, value=3: self.check_all(tab_idx, a, value))

                button_selectAll_layout.addWidget(arb_button_all)

                horizontalSpacer_3 = QSpacerItem(40, 20, QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Minimum)

                button_selectAll_layout.addItem(horizontalSpacer_3)

                scrollArea = QScrollArea(tab)
                scrollArea.setObjectName(u"scrollArea")
                scrollArea.setMinimumSize(QSize(0, 175))
                scrollArea.setMaximumSize(QSize(16777215, 175))
                scrollArea.setWidgetResizable(True)
                scrollAreaWidgetContents = QWidget()
                scrollAreaWidgetContents.setObjectName(u"scrollAreaWidgetContents")
                scrollAreaWidgetContents.setGeometry(QRect(0, 0, 1147, 185))

                gridLayout = QGridLayout(scrollAreaWidgetContents)
                gridLayout.setObjectName(u"gridLayout")

                timeSeriesFrame_V = QFrame(tab)
                timeSeriesFrame_V.setObjectName(u"timeSeriesFrame")
                timeSeriesFrame_V.setMinimumSize(QSize(0, 175))
                timeSeriesFrame_V.setFrameShape(QFrame.Shape.StyledPanel)
                timeSeriesFrame_V.setFrameShadow(QFrame.Shadow.Raised)
                playSound_V = QPushButton(timeSeriesFrame_V)
                playSound_V.setObjectName(u"playSound_V")
                playSound_V.setGeometry(QRect(10, 125, 33, 33))
                playSound_V.setStyleSheet(u"    color: rgb(255, 255, 255);\n"
                "    border-radius: 12.5px; /* Adjust to half of width/height */\n"
                "    border-style: outset;\n"
                "    background-color: rgb(0, 255, 0);\n"
                "    padding: 2px;\n"
                "    min-width: 25px; /* Ensures a round shape */\n"
                "    min-height: 25px; /* Equal width and height */\n"
                "	font-size: 28px;\n"
                "	font-weight: bold\n"
                "")
                playSound_V.setText(QCoreApplication.translate("Dialog", u"\u25b6", None))
                playSound_V.clicked.connect(lambda _, tab_idx=i, dir="V": self.on_play_button_clicked(tab_idx, dir))
                
                gridLayout_2.addWidget(timeSeriesFrame_V, 5, 0, 1, 1)

                timeSeriesFrame_L = QFrame(tab)
                timeSeriesFrame_L.setObjectName(u"timeSeriesFrame_3")
                timeSeriesFrame_L.setMinimumSize(QSize(0, 175))
                timeSeriesFrame_L.setFrameShape(QFrame.Shape.StyledPanel)
                timeSeriesFrame_L.setFrameShadow(QFrame.Shadow.Raised)
                playSound_L = QPushButton(timeSeriesFrame_L)
                playSound_L.setObjectName(u"playSound_L")
                playSound_L.setGeometry(QRect(10, 125, 33, 33))
                playSound_L.setStyleSheet(u"    color: rgb(255, 255, 255);\n"
                "    border-radius: 12.5px; /* Adjust to half of width/height */\n"
                "    border-style: outset;\n"
                "    background-color: rgb(0, 255, 0);\n"
                "    padding: 2px;\n"
                "    min-width: 25px; /* Ensures a round shape */\n"
                "    min-height: 25px; /* Equal width and height */\n"
                "	font-size: 28px;\n"
                "	font-weight: bold\n"
                "")
                playSound_L.setText(QCoreApplication.translate("Dialog", u"\u25b6", None))
                playSound_L.clicked.connect(lambda _, tab_idx=i, dir="L": self.on_play_button_clicked(tab_idx, dir))

                gridLayout_2.addWidget(timeSeriesFrame_L, 6, 0, 1, 1)

                timeSeriesFrame_T = QFrame(tab)
                timeSeriesFrame_T.setObjectName(u"timeSeriesFrame_2")
                timeSeriesFrame_T.setMinimumSize(QSize(0, 195))
                timeSeriesFrame_T.setFrameShape(QFrame.Shape.StyledPanel)
                timeSeriesFrame_T.setFrameShadow(QFrame.Shadow.Raised)
                playSound_T = QPushButton(timeSeriesFrame_T)
                playSound_T.setObjectName(u"playSound_T")
                playSound_T.setGeometry(QRect(10, 145, 33, 33))
                playSound_T.setStyleSheet(u"    color: rgb(255, 255, 255);\n"
                "    border-radius: 12.5px; /* Adjust to half of width/height */\n"
                "    border-style: outset;\n"
                "    background-color: rgb(0, 255, 0);\n"
                "    padding: 2px;\n"
                "    min-width: 25px; /* Ensures a round shape */\n"
                "    min-height: 25px; /* Equal width and height */\n"
                "	font-size: 28px;\n"
                "	font-weight: bold\n"
                "")
                playSound_T.setText(QCoreApplication.translate("Dialog", u"\u25b6", None))
                playSound_T.clicked.connect(lambda _, tab_idx=i, dir="T": self.on_play_button_clicked(tab_idx, dir))

                self.tabs[i] = {
                    "tab": tab,
                    "timeSeriesFrame_V": timeSeriesFrame_V,
                    "playSound_V": playSound_V,
                    "timeSeriesFrame_L": timeSeriesFrame_L,
                    "playSound_L": playSound_L,
                    "timeSeriesFrame_T": timeSeriesFrame_T,
                    "playSound_T": playSound_T,
                    "reel_button_all": reel_button_all,
                    "slag_button_all": slag_button_all,
                    "arb_button_all": arb_button_all,
                    "transients": {}
                }

                row = 0

                self.play_buttons = {}

                for j, transient in enumerate(transients):
                    
                    self.tabs[i]["transients"][j] = {}
                    

                    trans_temp = transient["transients"]

                    timestamp = search_json(trans_temp, "timestamp")[0]
                    date = search_json(trans_temp, "datetime")[0]                
                    labels = search_json(trans_temp, "label")
                    values = search_json(trans_temp, "value")
                    freqs = search_json(trans_temp, "frequency")

                    if 'type' in errors_list[timestamp]:
                        self.selected_vars[adresse][timestamp] = errors_list[timestamp]['type']
                    else:
                        self.selected_vars[adresse][timestamp] = 0

                    percentages = []
                    for value, freq in zip(values, freqs):
                        per = calculate_vibration_data(freq, 3, vibration=value, percentage=None)
                        
                        percentages.append(per)

                    gridLayout_3 = QGridLayout()
                    gridLayout_3.setObjectName("gridLayout_{}".format(timestamp))
                    gridLayout_3.setContentsMargins(-1, -1, -1, 0)
                    #horizontalSpacer_2 = QSpacerItem(15, 20, QSizePolicy.Policy.Minimum, QSizePolicy.Policy.Minimum)

                    #gridLayout_3.addItem(horizontalSpacer_2, row, 0, 1, 1)

                    showTransient_Button = QPushButton(scrollAreaWidgetContents)
                    showTransient_Button.setObjectName(u"pushButton_2")
                    sizePolicy2 = QSizePolicy(QSizePolicy.Policy.Fixed, QSizePolicy.Policy.Fixed)
                    sizePolicy2.setHorizontalStretch(0)
                    sizePolicy2.setVerticalStretch(0)
                    sizePolicy2.setHeightForWidth(showTransient_Button.sizePolicy().hasHeightForWidth())
                    showTransient_Button.setSizePolicy(sizePolicy2)
                    showTransient_Button.setMinimumSize(QSize(85, 35))

                    showTransient_Button.clicked.connect(lambda _, tab_idx=i, btn_idx=j, a=adresse, t=transient, dirs=labels, grid_layout=gridLayout_3: self.update_time_series(tab_idx, btn_idx, a, t, dirs, grid_layout))
                    
                    showTransient_Button.setText(QCoreApplication.translate("Dialog", u"Vis transient", None))

                    gridLayout_3.addWidget(showTransient_Button, row, 1, 1, 1)

                    transient_info_label_date = QLabel(scrollAreaWidgetContents)
                    transient_info_label_date.setObjectName(u"transient_info_label")
                    sizePolicy = QSizePolicy(QSizePolicy.Policy.Maximum, QSizePolicy.Policy.Maximum)
                    sizePolicy.setHorizontalStretch(100)
                    sizePolicy.setVerticalStretch(100)
                    sizePolicy.setHeightForWidth(transient_info_label_date.sizePolicy().hasHeightForWidth())
                    transient_info_label_date.setSizePolicy(sizePolicy)
                    transient_info_label_date.setMinimumSize(QSize(150, 0))
                    transient_info_label_date.setText(QCoreApplication.translate("Dialog", u"Dato: {}".format(date), None))

                    gridLayout_3.addWidget(transient_info_label_date, row, 4, 1, 1)

                    transient_info_label_data = QLabel(scrollAreaWidgetContents)
                    transient_info_label_data.setObjectName(u"label_3")
                    sizePolicy1 = QSizePolicy(QSizePolicy.Policy.Maximum, QSizePolicy.Policy.Preferred)
                    sizePolicy1.setHorizontalStretch(0)
                    sizePolicy1.setVerticalStretch(0)
                    sizePolicy1.setHeightForWidth(transient_info_label_data.sizePolicy().hasHeightForWidth())
                    transient_info_label_data.setSizePolicy(sizePolicy1)
                    transient_info_label_data.setMinimumSize(QSize(200, 0))

                    display_text = u"<html><head/><body><p>"
                    for label, value, freq, per in zip(labels, values, freqs, percentages):
                        display_text += f"{label}: {value} mm/s, {freq} Hz, {per}%<br/>"
                    
                    display_text = display_text[:-5] + "</p></body></html>" 

                    transient_info_label_data.setText(QCoreApplication.translate("Dialog", 
                                                                                 u"{}".format(display_text), None))

                    gridLayout_3.addWidget(transient_info_label_data, row, 5, 1, 1)

                    reel_checkbox = QCheckBox(scrollAreaWidgetContents)
                    reel_checkbox.setObjectName(u"reel_checkbox")
                    reel_checkbox.setMaximumSize(QSize(50, 16777215))        
                    reel_checkbox.setText(QCoreApplication.translate("Dialog", u"Reel", None))
                    reel_checkbox.setChecked(self.selected_vars[adresse][timestamp] == 1)
                    reel_checkbox.stateChanged.connect(
                        lambda state, a=adresse, t=timestamp, tab_idx=i, btn_idx=j, value=1: self.update_variable(state, a, t, tab_idx, btn_idx, value)
                    )
                    
                    self.tabs[i]['transients'][j]["reel_checkbox"] = reel_checkbox

                    gridLayout_3.addWidget(reel_checkbox, row, 6, 1, 1)

                    slag_checkbox = QCheckBox(scrollAreaWidgetContents)
                    slag_checkbox.setObjectName(u"slag_checkbox")
                    slag_checkbox.setMaximumSize(QSize(50, 16777215))
                    slag_checkbox.setText(QCoreApplication.translate("Dialog", u"Slag", None))
                    slag_checkbox.setChecked(self.selected_vars[adresse][timestamp] == 2)
                    slag_checkbox.stateChanged.connect(
                        lambda state, a=adresse, t=timestamp, tab_idx=i, btn_idx=j, value=2: self.update_variable(state, a, t, tab_idx, btn_idx, value)
                    )

                    self.tabs[i]['transients'][j]["slag_checkbox"] = slag_checkbox

                    gridLayout_3.addWidget(slag_checkbox, row, 7, 1, 1)

                    arbe_checkbox = QCheckBox(scrollAreaWidgetContents)
                    arbe_checkbox.setObjectName(u"arbe_checkbox")
                    arbe_checkbox.setMaximumSize(QSize(175, 16777215))
                    arbe_checkbox.setText(QCoreApplication.translate("Dialog", u"Arbejde t\u00e6t ved m\u00e5leren", None))
                    arbe_checkbox.setChecked(self.selected_vars[adresse][timestamp] == 3)
                    arbe_checkbox.stateChanged.connect(
                        lambda state, a=adresse, t=timestamp, tab_idx=i, btn_idx=j, value=3: self.update_variable(state, a, t, tab_idx, btn_idx, value)
                    )

                    self.tabs[i]['transients'][j]["arbe_checkbox"] = arbe_checkbox

                    gridLayout_3.addWidget(arbe_checkbox, row, 8, 1, 1)        
                    
                    line = QFrame(scrollAreaWidgetContents)
                    line.setObjectName(u"line")
                    line.setEnabled(True)
                    line.setMinimumSize(QSize(0, 0))
                    line.setStyleSheet(u"color: rgb(255, 255, 255)")
                    line.setFrameShadow(QFrame.Shadow.Plain)
                    line.setLineWidth(25)
                    line.setMidLineWidth(0)
                    line.setFrameShape(QFrame.Shape.VLine)

                    self.tabs[i]['transients'][j]["active_transient_indicator"] = line

                    gridLayout_3.addWidget(line, row, 9, 1, 1)

                    horizontalSpacer_2 = QSpacerItem(17, 20, QSizePolicy.Policy.Minimum, QSizePolicy.Policy.Minimum)

                    gridLayout_3.addItem(horizontalSpacer_2, row, 10, 1, 1)

                    horizontalSpacer = QSpacerItem(450, 50, QSizePolicy.Policy.Maximum, QSizePolicy.Policy.Minimum)

                    gridLayout_3.addItem(horizontalSpacer, row, 11, 1, 1)

                    gridLayout.addLayout(gridLayout_3, row, 0, 1, 1)

                    row += 1
                


                gridLayout_2.addWidget(timeSeriesFrame_T, 7, 0, 1, 1)



                scrollArea.setWidget(scrollAreaWidgetContents)
                gridLayout_2.addWidget(scrollArea, 4, 0, 1, 1)

                #self.tabWidget.addTab(self.tab_2, "")

            else:

                verticalLayout_4 = QVBoxLayout(tab)
                verticalLayout_4.setObjectName(u"verticalLayout_4")

                
                label_5 = QLabel(tab)
                label_5.setObjectName(u"label_5")
                label_5.setText(QCoreApplication.translate("Dialog", 
                                                            u"<html><head/><body><p align=\"center\"><span style=\" font-size:14pt; font-weight:700;\">Transienter for {}, {}.</span></p></body></html>".format(adresse, sensor_id), None))
                verticalLayout_4.addWidget(label_5)

                label_6 = QLabel(tab)
                label_6.setObjectName(u"label_6")
                label_6.setText(QCoreApplication.translate("Dialog", 
                u"<html><head/><body><p align=\"center\"><span style=\" font-size:12pt;\">Ingen transienter</span></p></body></html>", None))

                verticalLayout_4.addWidget(label_6)
                verticalLayoutWidget = QWidget(tab)
                verticalLayoutWidget.setObjectName(u"verticalLayoutWidget_2")
                verticalLayoutWidget.setGeometry(QRect(0, 20, 1391, 111))

                verticalSpacer_2 = QSpacerItem(20, 40, QSizePolicy.Policy.Minimum, QSizePolicy.Policy.Expanding)

                verticalLayout_4.addItem(verticalSpacer_2)

                verticalLayoutWidget_layout = QVBoxLayout(verticalLayoutWidget)
                verticalLayoutWidget_layout.setObjectName(u"verticalLayout_4")
                verticalLayoutWidget_layout.setContentsMargins(0, 0, 0, 0)

                label_6 = QLabel(verticalLayoutWidget)
                label_6.setObjectName(u"label_6")

                verticalLayoutWidget_layout.addWidget(label_6)

        self.verticalLayout_3.addWidget(self.tabWidget)
        self.tabWidget.setCurrentIndex(0)

    def cancel_MainWindow(self):
        self.reject()

    def check_all(self, tab_idx, adresse, value):
        
        for timestamp, val in self.selected_vars[adresse].items():
            val = value
        
        for btn_idx in self.tabs[tab_idx]['transients']:  # Iterate over all button indices
            if value == 1:
                self.tabs[tab_idx]['transients'][btn_idx]["reel_checkbox"].setChecked(True)
                self.tabs[tab_idx]['transients'][btn_idx]["slag_checkbox"].setChecked(False)
                self.tabs[tab_idx]['transients'][btn_idx]["arbe_checkbox"].setChecked(False)
            elif value == 2:
                self.tabs[tab_idx]['transients'][btn_idx]["reel_checkbox"].setChecked(False)
                self.tabs[tab_idx]['transients'][btn_idx]["slag_checkbox"].setChecked(True)
                self.tabs[tab_idx]['transients'][btn_idx]["arbe_checkbox"].setChecked(False)
            else:
                self.tabs[tab_idx]['transients'][btn_idx]["reel_checkbox"].setChecked(False)
                self.tabs[tab_idx]['transients'][btn_idx]["slag_checkbox"].setChecked(False)
                self.tabs[tab_idx]['transients'][btn_idx]["arbe_checkbox"].setChecked(True)

    def update_variable(self, state, adresse, timestamp, tab_idx, btn_idx, value):
        if state != 2:
            box_1 = self.tabs[tab_idx]['transients'][btn_idx]["reel_checkbox"].isChecked()
            box_2 = self.tabs[tab_idx]['transients'][btn_idx]["slag_checkbox"].isChecked()
            box_3 = self.tabs[tab_idx]['transients'][btn_idx]["arbe_checkbox"].isChecked()

            if all(box_state is False for box_state in [box_1, box_2, box_3]):
                self.selected_vars[adresse][timestamp] = None

            return
        
        if value == 1:
            self.tabs[tab_idx]['transients'][btn_idx]["slag_checkbox"].setChecked(False)
            self.tabs[tab_idx]['transients'][btn_idx]["arbe_checkbox"].setChecked(False)
        elif value == 2:
            self.tabs[tab_idx]['transients'][btn_idx]["reel_checkbox"].setChecked(False)
            self.tabs[tab_idx]['transients'][btn_idx]["arbe_checkbox"].setChecked(False)
        else:
            self.tabs[tab_idx]['transients'][btn_idx]["reel_checkbox"].setChecked(False)
            self.tabs[tab_idx]['transients'][btn_idx]["slag_checkbox"].setChecked(False)
        
        self.selected_vars[adresse][timestamp] = value

    def show_waveform(self, frame, tab_idx, btn_idx, transient, direction, adresse):
        from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
        import matplotlib.pyplot as plt
        import numpy as np

        if "active_transient" in self.tabs[tab_idx] and self.tabs[tab_idx]["active_transient"]:
            self.tabs[tab_idx]["active_transient"].setStyleSheet("color: rgb(255, 255, 255)")
        
        self.tabs[tab_idx]["active_transient"] = self.tabs[tab_idx]['transients'][btn_idx]["active_transient_indicator"]

        # Set the selected one to black
        self.tabs[tab_idx]["active_transient"].setStyleSheet("color: rgb(0, 0, 0)")


        def round_to_nearest(value):
            
            """Round value to nearest 10, 1, 0.1, or 0.01 based on magnitude."""
            if abs(value) >= 10:
                return np.ceil(value / 10) * 10
            elif abs(value) >= 1:
                return np.ceil(value)
            elif abs(value) >= 0.1:
                return np.ceil(value * 10) / 10
            else:
                return np.ceil(value * 100) / 100

        def plot_waveform(data, ax, direction):
            
            """Plot the waveform data."""
            plt.close()
            ax.clear()
            ax.plot(data['time'], data['value'], label=direction, linewidth=1)
            #ax.set_title('Transient tidsserie')
            #ax.set_xlabel('Tid [s]')

            ymin, ymax = ax.get_ylim()
            
            ymin_rounded = -round_to_nearest(abs(ymin))
            ymax_rounded = round_to_nearest(ymax)
            
            # Set new limits
            ax.set_ylim(ymin_rounded, ymax_rounded)

            # Generate nice y-ticks
            
            yticks = np.array([ymax_rounded, ymax_rounded/2, 0, ymin_rounded/2, ymin_rounded])

            if 0 not in yticks:
                yticks = np.append(yticks, 0)
                yticks = np.sort(yticks)

            # Set new y-ticks
            ax.set_yticks(yticks)
            
            if direction != "T":
                ax.set_xticklabels([])
            else:
                ax.set_xlabel('Tid [s]', labelpad = 20, fontsize=10)
                ax.xaxis.set_label_coords(1.03, -0.05)  # Adjust x and y position of label



            #yticks = np.linspace(ymin, ymax, num=6)  # Adjust `num` for more/less ticks
            #ax.set_yticks(yticks)

            ax.set_ylabel(f'{direction} [mm/s]', rotation=0, labelpad = 20, fontsize=10)
            ax.yaxis.set_label_coords(-0.09, 0.9)  # Adjust x and y position of label

             
            #Customize the spines to create a frame

            ax.grid()
            ax.legend()
            #plt.tight_layout()
            plt.draw()

        def load_waveform_data(file_path):
            import pandas as pd
            """Load waveform data from a CSV file."""
            if os.path.exists(file_path):
                return pd.read_csv(file_path, sep="\t")
            else:
                return pd.DataFrame(data={"time": [], "value": []})

        timestamp = search_json(transient, "timestamp")[0]
        file_path = f"{self.cache_main_dir}/{adresse}/Ascii/data_waveform/{timestamp}/{direction}_waveform.dat"  # Adjust the naming convention as necessary
        waveform_data = load_waveform_data(file_path)

        self.tabs[tab_idx][f"waveform_{direction}"] = waveform_data

        """ Embeds a Matplotlib figure inside the given frame """

        # Clear previous widgets in the frame
        if frame.layout() != None:
            for i in reversed(range(frame.layout().count())):
                frame.layout().itemAt(i).widget().setParent(None)

        fig, ax = plt.subplots(figsize=(frame.size().width()/100, ((frame.size().height()-50)/100)))
        
        plot_waveform(waveform_data, ax, direction)

        plt.tight_layout()

        canvas = FigureCanvas(fig)

        # Set layout for frame if not already set
        if not frame.layout():
            frame.setLayout(QVBoxLayout())

        frame.layout().addWidget(canvas)
        canvas.draw()

        frame.layout().setContentsMargins(0, 0, 0, 0)  # Add margin if necessary

    def update_time_series(self, tab_idx, btn_idx, adresse, transient, directions, grid_layout):   
        
        """ Updates time series frames based on the tab and button index """ 

        frames = self.tabs[tab_idx]

        for dir in directions:
            self.show_waveform(frames[f"timeSeriesFrame_{dir}"], tab_idx, btn_idx, transient, dir, adresse)

        frames["playSound_V"].raise_()
        frames["playSound_L"].raise_()
        frames["playSound_T"].raise_()

    def on_play_button_clicked(self, tab_idx, direction):
        import threading
        import numpy as np
        import sounddevice as sd
        def play_sound(waveform, sampling_rate):
            """Play the waveform sound."""
            sd.play(waveform, samplerate=sampling_rate)
            sd.wait()  # Wait until the sound finishes playing
        
        waveform_data = self.tabs[tab_idx][f"waveform_{direction}"]

        if waveform_data is not None:
            waveform = waveform_data["value"] / np.max(np.abs(waveform_data["value"]))
            timelength = waveform_data["time"].iloc[-1] - waveform_data["time"].iloc[0]
            sampling_rate = len(waveform_data) / timelength  # Calculating the sampling rate

            # Create a thread for sound playback
            playback_thread = threading.Thread(target=play_sound, args=(waveform, sampling_rate))
            playback_thread.start()  # Start the thread for sound playback

    def remove_checked_transients(self):
        import copy
        from openpyxl import Workbook, load_workbook
        from datetime import datetime

        def create_excel_sheet(wb, address, sheet_data, sensor_id, Reel_count, Slag_count, Arbe_count):  
            from openpyxl.styles import PatternFill

            fill_color = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

            # Parsing and organizing data for each address
            # Create a new sheet with the address name
            invalid_chars = ['\\', '/', '*', '[', ']', ':', '?']
            for char in invalid_chars:
                address = address.replace(char, "_")  # Replace invalid characters with an underscore

            # Create a sheet with the sanitized title
            ws = wb.create_sheet(title=address)

            # Setup headers and static values
            ws['A1'] = "Reel"
            ws['B1'] = Reel_count
            ws['A2'] = "Slag"
            ws['B2'] = Slag_count
            ws['A3'] = "Arbejde tæt på måleren"
            ws['B3'] = Arbe_count
            ws['A5'] = "Dato"
            ws['B5'] = "V [mm/s]"
            ws['C5'] = "L [mm/s]"
            ws['D5'] = "T [mm/s]"
            ws['E5'] = "rV [%]"
            ws['F5'] = "rL [%]"
            ws['G5'] = "rT [%]"
            ws['H5'] = "freqV [hZ]"
            ws['I5'] = "freqL [hZ]"
            ws['J5'] = "freqT [hZ]"
            ws['K5'] = "Type"

            cells = [ws['A3'], ws['B5'], ws['C5'], ws['D5'], ws['E5'], ws['F5'], ws['G5'], ws['H5'], ws['I5'], ws['J5'], ws['K5']]
            for i, cell in enumerate(cells):
                if i == 0 or i == len(cells) - 1:
                    ws.column_dimensions[cell.column_letter].width = len("Arbejde tæt på måleren") + 2
                else:
                    ws.column_dimensions[cell.column_letter].width = len("freqV [hZ]") + 2
                


            # Write the data rows
            for row_idx, row_data in enumerate(sheet_data, start=6):  # start from row 6 (1-based index)
                temp = row_data[sensor_id]["transients"]
                temp_to_value = {item['label']: item['value'] for item in temp}
                temp_to_frequency = {item['label']: item['frequency'] for item in temp}
                #Calculate percentage for each label
                temp_to_percentage = {
                    item['label']: calculate_vibration_data(
                        temp_to_frequency[item['label']],
                        3,
                        temp_to_value[item['label']],
                        percentage=None
                    )
                    for item in temp
                    }
                if row_data['type'] == 1:
                    transient_type = "Reel"
                elif row_data["type"] == 2:
                    transient_type = "Slag"
                else:
                    transient_type = "Arbejde tæt på måleren"

                ws[f'A{row_idx}'] = row_data['datetime']
                ws[f'B{row_idx}'] = float(temp_to_value.get('V', '')) if temp_to_value.get('V', '') != None else 0
                ws[f'C{row_idx}'] = float(temp_to_value.get('L', '')) if temp_to_value.get('L', '') != None else 0
                ws[f'D{row_idx}'] = float(temp_to_value.get('T', '')) if temp_to_value.get('T', '') != None else 0
                ws[f'E{row_idx}'] = float(temp_to_percentage.get('V', '')) if temp_to_percentage.get('V', '') != None else 0
                ws[f'F{row_idx}'] = float(temp_to_percentage.get('L', '')) if temp_to_percentage.get('L', '') != None else 0
                ws[f'G{row_idx}'] = float(temp_to_percentage.get('T', '')) if temp_to_percentage.get('T', '') != None else 0
                ws[f'H{row_idx}'] = float(temp_to_frequency.get('V', '')) if temp_to_frequency.get('V', '') != None else 0
                ws[f'I{row_idx}'] = float(temp_to_frequency.get('L', '')) if temp_to_frequency.get('L', '') != None else 0
                ws[f'J{row_idx}'] = float(temp_to_frequency.get('T', '')) if temp_to_frequency.get('T', '') != None else 0
                ws[f'K{row_idx}'] = transient_type

                if transient_type == "Slag":
                    for col_idx in range(1, 11):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        cell.fill = fill_color

        def append_excel_sheet(wb, adresse, sheet_data, sensor_id, project_id, today):        

            ws = wb["Traning Set"]

            # Read existing data into a set
            existing_rows = set()
            for row in ws.iter_rows(values_only=True):
                existing_rows.add(tuple(row))

            # Write the data rows

            new_rows = []

            for row_data in sheet_data:  # start from row 6 (1-based index)
                timestamp = row_data["timestamp"]
                datetime = row_data["datetime"]
                temp = row_data[sensor_id]["transients"]
                temp_to_value = {item['label']: item['value'] for item in temp}
                temp_to_frequency = {item['label']: item['frequency'] for item in temp}
                
                if row_data['type'] == 1:
                    transient_type = "Reel"
                elif row_data["type"] == 2:
                    transient_type = "Slag"
                else:
                    transient_type = "Arbejde tæt på måleren"

                new_row = (
                    str(self.project_id),
                    f"{self.atr} - {self.name}",
                    adresse,
                    sensor_id,
                    timestamp,
                    datetime,
                    today,
                    float(temp_to_value.get('V', 0)) if temp_to_value.get('V') is not None else 0,
                    float(temp_to_frequency.get('V', 0)) if temp_to_frequency.get('V') is not None else 0,
                    float(temp_to_value.get('L', 0)) if temp_to_value.get('L') is not None else 0,
                    float(temp_to_frequency.get('L', 0)) if temp_to_frequency.get('L') is not None else 0,
                    float(temp_to_value.get('T', 0)) if temp_to_value.get('T') is not None else 0,
                    float(temp_to_frequency.get('T', 0)) if temp_to_frequency.get('T') is not None else 0,
                    transient_type
                )

                # Check if an identical row exists (excluding columns G and N)
                identical_row = None
                for existing_row in existing_rows:
                    if (existing_row[:6] + existing_row[7:13] == new_row[:6] + new_row[7:13]):
                        identical_row = existing_row
                        break

                # If an identical row is found, remove it
                if identical_row:
                    if identical_row[13] != new_row[13]:
                        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
                            if tuple(cell.value for cell in row) == identical_row:
                                ws.delete_rows(row[0].row, 1)
                                break
                        new_rows.append(new_row)
                else:
                    # Append the new row
                    new_rows.append(new_row)

            for row in new_rows:
                ws.append(row)
        
        self.filename_transient = os.path.join(self.transient_data_folder, "transient_data.xlsx")
        wb_trans = load_workbook(self.filename_transient)

        today = datetime.today().strftime('%Y-%m-%d %H:%M:%S')

        wb = Workbook()
        wb.remove(wb['Sheet'])  # Remove the default sheet created by openpyxl

        filename = os.path.join(self.output_folder, "Transient_Resultater.xlsx")
        
        self.new_trans_dict = {}
        
        # Iterate through each adresse and its associated checkboxes
        for adresse in self.selected_vars:
            selected_vars_for_adresse = self.selected_vars[adresse] # Access the check_vars stored in each bottom frame
            
            reel_count = 0
            slag_count = 0
            arbe_count = 0

            above_limit_5 = False

            base_key = adresse.split("_")[0]

            if "Transienter" not in self.trans_dict[adresse].keys():
                self.trans_dict[adresse]["Transienter"] = []

            if self.trans_dict[adresse]["Transienter"] == []:
                self.trans_dict[adresse]["Reel"] = reel_count
                self.trans_dict[adresse]["Slag"] = slag_count
                self.trans_dict[adresse]["Arbe"] = arbe_count
                self.trans_dict[adresse]["above_limit_5"] = above_limit_5

                if not base_key in self.new_trans_dict.keys():
                    self.new_trans_dict[adresse] = copy.deepcopy(self.trans_dict[adresse])

                    self.new_trans_dict[adresse]["Reel"] = reel_count
                    self.new_trans_dict[adresse]["Slag"] = slag_count
                    self.new_trans_dict[adresse]["Arbe"] = arbe_count
                    self.new_trans_dict[adresse]["above_limit_5"] = above_limit_5
                else:
                    self.new_trans_dict[base_key]["Reel"] += reel_count
                    self.new_trans_dict[base_key]["Slag"] += slag_count
                    self.new_trans_dict[base_key]["Arbe"] += arbe_count
                    self.new_trans_dict[base_key]["above_limit_5"] = self.new_trans_dict[base_key]["above_limit_5"] if self.trans_dict[adresse]["above_limit_5"] == False else self.trans_dict[adresse]["above_limit_5"]

                continue
            errors = self.trans_dict[adresse]["Transienter"]

            # Iterate over each timestamp and update the error type based on checkbutton var
            for timestamp, var in selected_vars_for_adresse.items():
                error = next(error for error in errors if search_json(error, "timestamp")[0] == timestamp)
                error["type"] = var

                # Count the occurrences of each type (1 = Reel, 2 = Slag, 3 = Arbejde tæt ved måleren)
                try:
                    if error["type"] == 1:
                        reel_count += 1
                    elif error["type"] == 2:
                        slag_count += 1
                    elif error["type"] == 3:
                        arbe_count += 1
                    else:
                        raise ValueError(f"Manglende analyse af transient for {adresse}")
                except ValueError:
                    self.logger.log(f"Mangler analyse af transient(er) for {adresse}!",2)
                    QMessageBox.warning(None, f"{adresse}", f"Mangler analyse af transient(er) for {adresse}")
                    return
                    
                    
                if error["type"] != 2 and self.info_dict[adresse]["limit"] == 3:
                    values = search_json(error, "value")
                    freqs = search_json(error, "frequency")
                    per_5 = []
                    for value, freq in zip(values, freqs):
                        per = calculate_vibration_data(freq, 5, vibration=value, percentage=None)
                        per_5.append(per)
                    if any(per>100 for per in per_5):
                        error["above_limit_5"] = True
                        above_limit_5 = True
                    else:
                        error["above_limit_5"] = False
                else:
                    error["above_limit_5"] = False
            
            self.trans_dict_adresse = os.path.join(self.cache_main_dir, f"{adresse}", "JSON", "data_Transienter.json")
            save_json(errors, self.trans_dict_adresse)

            self.trans_dict[adresse]["Reel"] = reel_count
            self.trans_dict[adresse]["Slag"] = slag_count
            self.trans_dict[adresse]["Arbe"] = arbe_count
            self.trans_dict[adresse]["above_limit_5"] = above_limit_5

            if base_key in self.new_trans_dict.keys():
                self.new_trans_dict[base_key]["Reel"] += reel_count
                self.new_trans_dict[base_key]["Slag"] += slag_count
                self.new_trans_dict[base_key]["Arbe"] += arbe_count
                self.new_trans_dict[base_key]["above_limit_5"] = self.new_trans_dict[base_key]["above_limit_5"] if self.trans_dict[adresse]["above_limit_5"] == False else self.trans_dict[adresse]["above_limit_5"]

                self.new_trans_dict[base_key]["Transienter"] += self.trans_dict[adresse]["Transienter"]
                self.new_trans_dict[base_key]["sensor_id"] =  [self.new_trans_dict[base_key]["sensor_id"], self.trans_dict[adresse]["sensor_id"]]
            else:
                self.new_trans_dict[adresse] = copy.deepcopy(self.trans_dict[adresse])


        save_json(self.new_trans_dict, self.trans_dict_final_path)
        save_json(self.trans_dict, self.trans_dict_path)            

        for adress, data in self.trans_dict.items():
            sensor_id = search_json(self.trans_dict[adress], "sensor_id")[0]
            Reel_count = data["Reel"]
            Slag_count = data["Slag"]
            Arbe_count = data["Arbe"]

            # Create Excel sheet for the address
            create_excel_sheet(wb, adress, data["Transienter"], sensor_id, Reel_count, Slag_count, Arbe_count)
            if data["Transienter"] != []:
                append_excel_sheet(wb_trans, adress, data["Transienter"], sensor_id, self.project_id, today)

        wb.save(filename)
        wb_trans.save(self.filename_transient)

        self.accept()

    def retranslateUi(self, Form):
        Form.setWindowTitle(QCoreApplication.translate("Dialog", u"Transient", None))
        
        self.cancelButton.setText(QCoreApplication.translate("Dialog", u"Afbryd", None))
        self.continueButton.setText(QCoreApplication.translate("Dialog", u"Forts\u00e6t", None))
        #self.label_4.setText(QCoreApplication.translate("MainWindow", u"<html><head/><body><p align=\"center\">TextLabel</p></body></html>", None))
    # retranslateUi


#if __name__ == "__main__":
#    project_id = "108790"
#
#    atr = "A223276-087"
#    name = "Hadsundvej Syd, Gistrup"
#
#    cache_main_dir = "O:\\A000000\\A004371\\3_Pdoc\\Afrapportering\\test\\A223276-087 - Hadsundvej Syd, Gistrup"
#    app = QtWidgets.QApplication(sys.argv)
#
#    ui = TransientUI(project_id, atr, name, cache_main_dir, "", "O:\\A000000\\A004371\\3_Pdoc\\Afrapportering\\test")
#
#    ui.show()
#
#    app.exec()

